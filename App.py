
import openpyxl
from datetime import datetime
import os
from tkinter import filedialog
import tkinter as tk
import subprocess

def browse_file():
    file_path = filedialog.askopenfilename()
    return file_path

def process_excel(file_path):
    if not file_path:
        print("No file selected.")
        return

    # Load the Excel workbook
    workbook = openpyxl.load_workbook(file_path)

    # Select the active sheet
    sheet = workbook.active

    # Create a list to store the row numbers to delete
    rows_to_delete = []

    # Iterate over all rows in column A starting from row 2
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True), start=2):
        print(row, row_number)
        # Check if the cell value is not None
        if row[0] is not None:
            try:
                # Attempt to parse the cell value as a date
                date_value = datetime.strptime(row[0], "%d/%m/%Y")
                # If successful, replace the cell value with the datetime object
                sheet.cell(row=row_number, column=1, value=date_value)
            except ValueError:
                # If parsing fails, mark the row for deletion
                rows_to_delete.append(row_number)
        else:
            # If the cell value is None, mark the row for deletion
            rows_to_delete.append(row_number)

    # Delete rows marked for deletion (starting from the last row to avoid shifting issues)
    for row_number in reversed(rows_to_delete):
        sheet.delete_rows(row_number)

    # Add header row
    sheet.insert_rows(1)
    # Set column headers
    sheet.cell(row=1, column=1, value="Date")
    sheet.cell(row=1, column=2, value="Description")
    sheet.cell(row=1, column=3, value="Amount")
    sheet.delete_rows(2)

    # Save the modified workbook with a fixed name in the project folder
    project_directory = os.getcwd()  # Get current working directory
    modified_file_path = os.path.join(project_directory, "modified_excel_file.xlsx")
    workbook.save(modified_file_path)
    print(f"Modified Excel file saved as: {modified_file_path}")

    # Run other scripts
    script_paths = [
        "pip_install0.py",
        "Convert_json-2.py",
        "checkdescription3.py",
        "orginzebyflag4.py",
        "flagonevalue5.py",
        "delete line output4-6.py",
        "pushtogit7.py",
        "updatesheets-8.py",
        "Sum9.py"
    ]

    for script_path in script_paths:
        subprocess.run(["python", script_path, modified_file_path])

if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()  # Hide the main window

    file_path = browse_file()
    if file_path:
        process_excel(file_path)
